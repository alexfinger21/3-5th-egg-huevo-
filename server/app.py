from openpyxl import load_workbook
import os

from flask import *

app = Flask(__name__)

if __name__ == '__main__':
   app.run()

@app.route('/')
def home():
   return render_template('index.html')
   
@app.route('/autocomplete', methods=["GET", "POST"])
def autocomplete():
   print(request.method)
   if request.method == "POST":
      req = request.get_json()
      if req["type"] == "state": #autocomplete for state
         options = []
         
         for k, state in state_abbrev.items():
               options.append(state)
         
         return jsonify(options)
      elif req["type"] == "county":
         return jsonify(counties_and_cities[state_unabbrev[req["state"]]])
      elif req["type"] == "data":
         stuff = ""
         county_no = req["county"][:-4]
         state = req["state"]
         if state in state_abbrev.values():
            state = state_unabbrev[req["state"]]
         actual_location = ""
         if county_no not in counties_list: 
            for county in counties_in_states[state]:
               for city in cities_in_counties[county]:
                  if county_no == city:
                     actual_location = county
                     break
         else:  
               actual_location = county_no
         actual_location += ", " + state
         
         stuff = county_no + ", " + state + ("\n\nPopulation: {population}\nMedian House Price: ${median_house_price}\nMedian Rent Price: ${median_rent_price}\n").format(population=county_dictionary[actual_location]["Population"], median_house_price=county_dictionary[actual_location]["Median House Price"], median_rent_price=county_dictionary[actual_location]["Median Rent Price"]) 

         #col
         if county_no in cost_of_living.keys():
            stuff += "Cost Of Living: $" + cost_of_living[county_no]

         if county_no in state_crime_dictionary.keys():
            state_crime_dictionary = {}
            state_crime_spreadsheet = state_crime_abbrev['OH']
            for i in range(6, 292):
               if state_crime_spreadsheet.cell(row=i, column=2).value == county_no:
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value] = {}
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Total Offences"] = state_crime_spreadsheet.cell(row=i,column=3).value
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Assault Offences"] = state_crime_spreadsheet.cell(row=i,column=8).value
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Homicide Offences"] = state_crime_spreadsheet.cell(row=i,column=12).value
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Kidnapping/Abduction"] = state_crime_spreadsheet.cell(row=i,column=19).value
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Burglary/Breaking and Entering"] = state_crime_spreadsheet.cell(row=i,column=29).value
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Fraud Offences"] = state_crime_spreadsheet.cell(row=i,column=34).value
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Identity Theft"] = state_crime_spreadsheet.cell(row=i,column=40).value
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Narcotic Offenses"] = state_crime_spreadsheet.cell(row=i,column=55).value
                  state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Crime Rate"] = state_crime_spreadsheet.cell(row=i,column=4).value / state_crime_spreadsheet.cell(row=i,column=3).value
                  
                  stuff += [state_crime_dictionary[req["county"]]["Assault Offences"], state_crime_dictionary[req["county"]]["Homicide Offences"], state_crime_dictionary[req["county"]]["Kidnapping/Abduction"], state_crime_dictionary[req["county"]]["Burglary/Breaking and Entering"], state_crime_dictionary[req["county"]]["Fraud Offences"], state_crime_dictionary[req["county"]]["Identity Theft"], state_crime_dictionary[req["county"]]["Narcotic Offenses"]]

                  stuff += [state_crime_dictionary[req["county"]]["Total Offences"], state_crime_dictionary[req["county"]]["Crime Rate"]]
                  #total offenses, crime rate
                  break
         return stuff
   else:
      return render_template('index.html')

state_abbrev = {
    'AK': 'Alaska',
    'AL': 'Alabama',
    'AR': 'Arkansas',
    'AZ': 'Arizona',
    'CA': 'California',
    'CO': 'Colorado',
    'CT': 'Connecticut',
    'DC': 'District of Columbia',
    'DE': 'Delaware',
    'FL': 'Florida',
    'GA': 'Georgia',
    'HI': 'Hawaii',
    'IA': 'Iowa',
    'ID': 'Idaho',
    'IL': 'Illinois',
    'IN': 'Indiana',
    'KS': 'Kansas',
    'KY': 'Kentucky',
    'LA': 'Louisiana',
    'MA': 'Massachusetts',
    'MD': 'Maryland',
    'ME': 'Maine',
    'MI': 'Michigan',
    'MN': 'Minnesota',
    'MO': 'Missouri',
    'MS': 'Mississippi',
    'MT': 'Montana',
    'NC': 'North Carolina',
    'ND': 'North Dakota',
    'NE': 'Nebraska',
    'NH': 'New Hampshire',
    'NJ': 'New Jersey',
    'NM': 'New Mexico',
    'NV': 'Nevada',
    'NY': 'New York',
    'OH': 'Ohio',
    'OK': 'Oklahoma',
    'OR': 'Oregon',
    'PA': 'Pennsylvania',
    'RI': 'Rhode Island',
    'SC': 'South Carolina',
    'SD': 'South Dakota',
    'TN': 'Tennessee',
    'TX': 'Texas',
    'UT': 'Utah',
    'VA': 'Virginia',
    'VT': 'Vermont',
    'WA': 'Washington',
    'WI': 'Wisconsin',
    'WV': 'West Virginia',
    'WY': 'Wyoming'
}
state_unabbrev = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "Hawaii": "HI",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "District of Columbia": "DC"}

#crime rate
state_crime_abbrev = {
   #  'AK': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Alaska_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'AL': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Alabama_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'AR': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Arkansas_Offense_Type_by_Agency_2021.xlxs")).active,
   #  'AZ': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Arizona_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'CA': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/California_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'CO': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Colorado_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'CT': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Connecticut_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'DC': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/District_of_Columbia_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'DE': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Delaware_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'FL': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Florida_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'GA': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Georgia_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'HI': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Hawaii_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'IA': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Iowa_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'ID': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Idaho_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'IL': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Illinois_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'IN': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Indiana_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'KS': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Kansas_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'KY': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Kentucky_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'LA': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Loisiana_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'MA': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Massachusetts_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'MD': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Maryland_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'ME': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Maine_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'MI': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Michigan_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'MN': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Minnesota_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'MO': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Missouri_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'MS': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Mississippi_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'MT': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Montana_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'NC': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/North_Carolina_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'ND': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/North_Dakota_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'NE': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Nebraska_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'NH': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/New_Hampshire_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'NJ': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/New_Jersey_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'NM': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/New_Mexico_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'NV': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Nevada_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'NY': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/New_York_Offense_Type_by_Agency_2021.xlsx")).active,
    'OH': load_workbook(os.path.abspath("Ohio_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'OK': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Oklahoma_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'OR': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Oregon_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'PA': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Pennsylvania_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'RI': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Rhode_Island_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'SC': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/South_Carolina_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'SD': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/South_Dakota_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'TN': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Tennessee_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'TX': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Texas_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'UT': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Utah_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'VA': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Virginia_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'VT': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Vermont_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'WA': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Washington_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'WI': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Wisconsin_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'WV': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/West_Virginia_Offense_Type_by_Agency_2021.xlsx")).active,
   #  'WY': load_workbook(os.path.abspath("server/State Tables Offenses by Agency/Wyoming_Offense_Type_by_Agency_2021.xlsx")).active
    
}

#rent prices
median_rent_prices_of_counties_spreadsheet = load_workbook(os.path.abspath("FY2023_FMR_50_county.xlsx")).active
global county_dictionary 
county_dictionary = {}
for i in range(2, 4766):
   x = median_rent_prices_of_counties_spreadsheet.cell(row=i, column=4).value + median_rent_prices_of_counties_spreadsheet.cell(row=i, column=6).value[(median_rent_prices_of_counties_spreadsheet.cell(row=i, column=6).value).find(","):(median_rent_prices_of_counties_spreadsheet.cell(row=i, column=6).value).find(",")+4]
   if ("County" in median_rent_prices_of_counties_spreadsheet.cell(row=i, column=4).value):
      if not x in county_dictionary.keys():
         county_dictionary[x] = {}
      county_dictionary[x]["Median Rent Price"] = median_rent_prices_of_counties_spreadsheet.cell(row=i, column = 8).value

#house prices
median_house_prices_of_counties_spreadsheet = load_workbook(os.path.abspath("2022-q3-county-median-prices-and-monthly-mortgage-payment-by-price-12-20-2022.xlsx")).active
for i in range(2, 3119):
   x = median_house_prices_of_counties_spreadsheet.cell(row=i, column = 1).value[:median_house_prices_of_counties_spreadsheet.cell(row=i, column = 1).value.find(",")+2] + state_unabbrev[median_house_prices_of_counties_spreadsheet.cell(row=i, column = 1).value[median_house_prices_of_counties_spreadsheet.cell(row=i, column = 1).value.find(",")+3:]]
   if ("County" in median_house_prices_of_counties_spreadsheet.cell(row=i, column=1).value):
      if not x in county_dictionary.keys():
         county_dictionary[x] = {}
      county_dictionary[x]["Median House Price"] = median_house_prices_of_counties_spreadsheet.cell(row=i, column = 2).value

#population
county_population_spreadsheet = load_workbook(os.path.abspath("co-est2021-pop.xlsx")).active
for i in range(6, 3148):
   x = county_population_spreadsheet.cell(row=i, column = 1).value[:county_population_spreadsheet.cell(row=i, column = 1).value.find(",")+2] + state_unabbrev[county_population_spreadsheet.cell(row=i, column = 1).value[county_population_spreadsheet.cell(row=i, column = 1).value.find(",")+2:]]
   x = x[1:]
   if ("County" in county_population_spreadsheet.cell(row=i, column=1).value):
      if not x in county_dictionary.keys():
         county_dictionary[x] = {}
      county_dictionary[x]["Population"] = county_population_spreadsheet.cell(row=i, column = 4).value
      
#cost of living
col_spreadsheet = load_workbook(os.path.abspath("advisorsmith_cost_of_living_index.xlsx")).active
cost_of_living =  {}
for i in range(2, 512):
   cost_of_living[col_spreadsheet.cell(row=i, column = 1).value] = (col_spreadsheet.cell(row=i, column=3).value)*0.01*1746

#county
counties_in_states = {}
counties_spreadsheet = load_workbook(os.path.abspath("list-counties-us-436j.xlsx")).active
for i in range(2, 3145):
   if "City" not in counties_spreadsheet.cell(row=i, column=2).value:
      if not state_unabbrev[counties_spreadsheet.cell(row=i, column=3).value] in counties_in_states.keys(): 
         counties_in_states[state_unabbrev[counties_spreadsheet.cell(row=i, column=3).value]] = []
      counties_in_states[state_unabbrev[counties_spreadsheet.cell(row=i, column=3).value]].append(counties_spreadsheet.cell(row=i, column=2).value)

#city
cities_in_states = {}
cities_spreadsheet = load_workbook(os.path.abspath("uscities.xlsx")).active
for i in range(2, 30411):
   if not cities_spreadsheet.cell(row=i, column=3).value in cities_in_states.keys(): 
      cities_in_states[cities_spreadsheet.cell(row=i, column=3).value] = []
   cities_in_states[cities_spreadsheet.cell(row=i, column=3).value].append(cities_spreadsheet.cell(row=i, column=1).value)
cities_in_states.pop("MA")

counties_in_states.pop("MA")
counties_and_cities = {}  
for key in counties_in_states.keys():
   counties_and_cities[key] = counties_in_states[key] + cities_in_states[key]
   
#cities in counties
cities_in_counties = {}
for i in range(2, 30411):
   if not (cities_spreadsheet.cell(row=i, column=6).value + " County") in cities_in_counties.keys(): 
      cities_in_counties[cities_spreadsheet.cell(row=i, column=6).value + " County"] = []
   cities_in_counties[cities_spreadsheet.cell(row=i, column=6).value + " County"].append(cities_spreadsheet.cell(row=i, column=1).value)

cities_list = []
counties_list = []
for i in range(2, 30411):
   cities_list.append(cities_spreadsheet.cell(row=i, column=1).value)
for i in range(2, 3145):
   counties_list.append(counties_spreadsheet.cell(row=i, column=2).value)

state_crime_dictionary = {}
state_crime_spreadsheet = state_crime_abbrev['OH']
for i in range(6, 292):
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value] = {}
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Total Offences"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=3).value
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Assault Offences"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=8).value
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Homicide Offences"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=12).value
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Kidnapping/Abduction"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=19).value
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Burglary/Breaking and Entering"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=29).value
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Fraud Offences"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=34).value
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Identity Theft"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=40).value
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Narcotic Offenses"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=55).value
      state_crime_dictionary[state_crime_spreadsheet.cell(row=i, column=2).value]["Crime Rate"] = median_house_prices_of_counties_spreadsheet.cell(row=i,column=3).value / median_house_prices_of_counties_spreadsheet.cell(row=i,column=2).value